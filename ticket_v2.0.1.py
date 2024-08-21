from tkinter import *
from tkinter import StringVar
import os
import tkinter.messagebox as msgbox
import tkinter.ttk
import tkinter.font
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
from pandastable import Table, TableModel
import re
from datetime import datetime
import json

# Define the font properties
font = Font(name='굴림체', size=9, bold=False)
nm_font = Font(name='맑은 고딕', size=11, bold=False)

with open('user_data_int.json', 'r', encoding='utf-8') as file:
    user_data_int = json.load(file)

with open('user_data_str.json', 'r', encoding='utf-8') as file:
    user_data_str = json.load(file)

class TicketNumber():
    def __init__(self, ticket_number):
        self._ticket_number = ticket_number

    @property
    def ticket_number(self):
        return self._ticket_number

    @ticket_number.setter
    def ticket_number(self, value):
        self._ticket_number = value
        ticket_var.set(self._ticket_number-1)  # Update the StringVar

class NonMember():
    def __init__(self, non_member_row):
        self.non_member_row = non_member_row

count = TicketNumber(1)
nm_row = NonMember(2)

def excel_loader():
    # Retrieve month and date
    month = f'{datetime.now().month:02}'
    date = f'{datetime.now().day:02}'

    file_name = f'전체이용자(편집)_{month}_{date}.xlsx'

    if os.path.exists(file_name):
        print('file found')
        return 
    else:
        workbook = load_workbook('전체이용자(편집).xlsx')
        workbook.save(file_name)
        print('file created')
        return

def jsonLoader(mont, day):
    # Retrieve month and date
    date = f'{day:02}'
    month = f'{mont:02}'

    # Set filename
    file_name = f'식권_판매정보_{month}_{date}'

    try:
        with open(f'{file_name}.json', 'r', encoding='utf-8') as file:
            data = json.load(file)
            if len(data) != 0:
                for d in data:
                    treeview.insert('', '0', text='', values=(f'{d['user_id']} {d['user_name']} {d['status']} {d['price']} {d['ticket_number']}'))
            return file_name
    except:
        with open(f'{file_name}.json', 'w', encoding='utf-8') as file:
            json.dump([], file)
        return file_name


def warn():
    msgbox.showwarning("경고", "회원번호 불일치")

def output(text):
    msgbox.showinfo("Output", f'{text}')

tk = Tk()
tk.title("식권 판매 프로그램")
tk.geometry("800x800")

# Create the DataFrame viewer at the top
frame = Frame(tk)
frame.pack(fill=BOTH, expand=True)

# Define the font for the Treeview
treeview_font = tkinter.font.Font(family="Arial", size=13)

style = tkinter.ttk.Style()
style.configure("Treeview", font=treeview_font)  # Set the font for the items
style.configure("Treeview.Heading", font=treeview_font)  # Set the font for the headings

treeview = tkinter.ttk.Treeview(
    frame,            
    columns=['UserID', 'Name', 'Status', 'price', 'ticketNumber'],
    displaycolumns=['UserID', 'Name', 'Status', 'price', 'ticketNumber'], show='headings')
treeview.pack(side=TOP, fill=BOTH, expand=True, padx=10, pady=10)

treeview.column("UserID", width=100, anchor="center")
treeview.heading("UserID", text="회원번호", anchor="center")

treeview.column("Name", width=100, anchor="center")
treeview.heading("Name", text="이름", anchor="center")

treeview.column("Status", width=150, anchor="center")
treeview.heading("Status", text="생활구분", anchor="center")

treeview.column("price", width=100, anchor="center")
treeview.heading("price", text="가격", anchor="center")

treeview.column("ticketNumber", width=100, anchor="center")
treeview.heading("ticketNumber", text="식권", anchor="center")

# Create a frame to hold the entry and buttons horizontally
input_frame = Frame(tk)
input_frame.pack(pady=10)

entry = Entry(input_frame)
entry.grid(row=0, column=0, padx=5)

_ = excel_loader()
jsonName = jsonLoader(datetime.now().month, datetime.now().day)

def search(event=None): # Add event parameter to handle binding
    input = entry.get()
    if any(char.isdigit() for char in input) == True:
        jsonSellTicketInt(input)
    else:
        jsonSellTicketStr(input)
    entry.delete(0, 'end')


def add_user(userID, name, status):
    # Assign appropriate price for users
    re_val = re.sub(r'\s+', '', status)
    match re_val:
        case '기초생활수급권자' | '차상위(저소득)':
            price = 0
        case '기타' | '국가유공자':
            price = 1750
        case '일반':
            price = 3500
        case _:
            price = 9999

    # Add new user to lists
    user_data_int[userID] = (name, status, price)
    user_data_str[name] = (userID, status, price)

    # Save those lists into JSONs
    with open('user_data_int.json', 'w', encoding='utf-8') as file:
        json.dump(user_data_int, file, ensure_ascii=False)
    with open('user_data_str.json', 'w', encoding='utf-8') as file:
        json.dump(user_data_str, file, ensure_ascii=False)

    saveNewUser(userID, name, status)

def add_non_member(name, reason_for_visit):
    jsonUpdate('비회원', name, '비회원', 3500)

# Function to delete the selected item
def refund():
    global jsonName
    selected_item = treeview.selection()  # Get selected item
    if selected_item:
        item = treeview.item(selected_item, 'values')
        userID = item[0]
        ticketNumber = int(item[4])
        count.ticket_number = ticketNumber
        deleteJSON(userID)
        treeview.delete(selected_item)  # Delete the selected item

def deleteJSON(userID):
    global jsonName
    with open(f'{jsonName}.json', 'r', encoding='utf-8') as file:
            user_data = json.load(file)

    # Creates a new list that only includes data that does not have matching userID    
    user_data = [data for data in user_data if data['user_id'] != userID]

    # Write the updated data back to the file
    with open(f'{jsonName}.json', 'w', encoding='utf-8') as file:
        json.dump(user_data, file, indent=4, ensure_ascii=False)


def jsonSellTicketInt(userID):
    global jsonName
    data = user_data_int.get(userID)
    if bool(data) == True:
        name = data[0]
        status = data[1]
        price = data[2]
        
        jsonUpdate(userID, name, status, price)
    else:
        warn()

def jsonSellTicketStr(userName):
    global jsonName
    data = user_data_str.get(userName)
    if bool(data) == True:
        userID = data[0]
        status = data[1]
        price = data[2]

        jsonUpdate(userID, userName, status, price)
    else:
        warn()

def jsonUpdate(userID, name, status, price):
    global jsonName
    new_data = {
            "user_id": userID,
            "user_name": name,
            "status": status,
            "price": price,
            "ticket_number": count.ticket_number 
        }

    treeview.insert('', '0', text='', values=(f'{userID} {name} {status} {price} {count.ticket_number}'))

    
    
    # Open the JSON file and load the data
    with open(f'{jsonName}.json', 'r', encoding='utf-8') as file:
        data = json.load(file)

    # Append new data to the list
    data.append(new_data)

    # Write the updated data back to the file
    with open(f'{jsonName}.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, indent=4, ensure_ascii=False)

    if count.ticket_number % 5 == 0:
        jsonBackUp()
    # Update Ticket Number
    count.ticket_number += 1

def reset_ticket_number():
    reset_value = int(entry.get())
    count.ticket_number = reset_value

def jsonBackUp():
    global jsonName

    with open(f'{jsonName}.json', 'r', encoding='utf-8') as file:
        data = json.load(file)

    jsonBack = f'{jsonName}_{count.ticket_number}.json'
    with open(jsonBack, 'w', encoding='utf-8'):
        json.dump(data, file, indent=4, ensure_ascii=False)


def saveNewUser(userID, name, status):
    workbook = load_workbook('전체이용자(편집).xlsx')
    sheet = workbook['회원']
    last_row = sheet.max_row + 1
    sheet[f'A{last_row}'].value = userID
    sheet[f'B{last_row}'].value = name
    sheet[f'C{last_row}'].value = status

    # Change font
    sheet[f'A{last_row}'].font = font
    sheet[f'B{last_row}'].font = font
    sheet[f'C{last_row}'].font = font

    # Align texts to center
    sheet[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'B{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'C{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'D{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'E{last_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Save workbook
    workbook.save('전체이용자(편집).xlsx')

def save_file():
    month = f'{datetime.now().month:02}'
    date = f'{datetime.now().day:02}'

    global jsonName
    file_name = f'전체이용자(편집)_{month}_{date}'
    dir = f'C:\\Users\\Administrator\\Desktop\\식권2024\\8월\\{file_name}.xlsx'
    # Define the font properties
    font = Font(name='굴림체', size=9, bold=False)
    nm_font = Font(name='맑은 고딕', size=11, bold=False)

    workbook = load_workbook('전체이용자(편집).xlsx')
    sheet = workbook['회원']
    nonSheet = workbook['비회원']

    with open(f'{jsonName}.json', 'r', encoding='utf-8') as file:
        user_data = json.load(file)

    for data in user_data:
        userID = data['user_id']
        name = data['user_name']
        status = data['status']
        price = int(data['price'])
        ticketNumber = int(data['ticket_number'])

        if str(userID) == '비회원':
            nonSheet[f'A{nm_row.non_member_row}'].value = name
            nonSheet[f'B{nm_row.non_member_row}'].value = status
            nonSheet[f'C{nm_row.non_member_row}'].value = price
            nonSheet[f'D{nm_row.non_member_row}'].value = ticketNumber
            nonSheet[f'A{nm_row.non_member_row}'].font = nm_font
            nonSheet[f'B{nm_row.non_member_row}'].font = nm_font
            nonSheet[f'C{nm_row.non_member_row}'].font = nm_font
            nonSheet[f'D{nm_row.non_member_row}'].font = nm_font
            nm_row.non_member_row += 1
            continue
        for cell in sheet['A']:
            if str(cell.value) == str(userID):
                sheet[f'D{cell.row}'].value = price
                sheet[f'E{cell.row}'].value = ticketNumber
                sheet[f'D{cell.row}'].font = font
                sheet[f'E{cell.row}'].font = font
    
    # Save workbook
    workbook.save(f'{file_name}.json')
    workbook.save(dir)

def addUserPopUp():
    addUserWindow = Toplevel(tk)
    addUserWindow.geometry('500x300')
    addUserWindow.title('신규회원추가')
    userID_text = Entry(addUserWindow)
    name_text = Entry(addUserWindow)
    status_text = Entry(addUserWindow)
    
    Label(addUserWindow, text='회원번호').grid(row=0, column=0, padx=5)
    userID_text.grid(row=0, column=1, padx=5)
    Label(addUserWindow, text='이름').grid(row=1, column=0, padx=5)
    name_text.grid(row=1, column=1, padx=5)
    Label(addUserWindow, text='생활구분').grid(row=2, column=0, padx=5)
    status_text.grid(row=2, column=1, padx=5)

    addUserButton = Button(addUserWindow, text='신규회원추가', command=lambda: add_user(userID_text.get(), name_text.get(), status_text.get()))
    addUserButton.grid(row=3, column=2, padx=5)

def nonUserPopUp():
    nonUserWindow = Toplevel(tk)
    nonUserWindow.geometry('500x300')
    nonUserWindow.title('비회원 판매')

    name_text = Entry(nonUserWindow)
    reason_for_visit_text = Entry(nonUserWindow)

    Label(nonUserWindow, text='이름').grid(row=0, column=0, padx=5)
    name_text.grid(row=0, column=1, padx=5)
    Label(nonUserWindow, text='방문용무').grid(row=1, column=0, padx=5)
    reason_for_visit_text.grid(row=1, column=1, padx=5)

    addNonUserButton = Button(nonUserWindow, text='비회원 판매', command=lambda: add_non_member(name_text.get(), reason_for_visit_text.get()))
    addNonUserButton.grid(row=2, column=2, padx=5)

# Create the StringVar after initializing the root window
ticket_var = StringVar()

search_button = Button(input_frame, text='검색', command=search)
search_button.grid(row=0, column=1, padx=5)

delete_button = Button(input_frame, text='환불', command=refund)
delete_button.grid(row=0, column=2, padx=5)

add_user_button = Button(input_frame, text='신규회원추가', command=addUserPopUp)
add_user_button.grid(row=0, column=3, padx=5)

reset_number = Button(input_frame, text='식권번호리셋', command=reset_ticket_number)
reset_number.grid(row=0, column=4, padx=5)

non_memeber_button = Button(input_frame, text='비회원', command=nonUserPopUp)
non_memeber_button.grid(row=0, column=5, padx=5)

save_button = Button(input_frame, text='저장', command=save_file)
save_button.grid(row=0, column=6, padx=5)

total_tickets_sold = Label(input_frame, textvariable=ticket_var, font=('Helvetica', 25, 'bold'))
total_tickets_sold.grid(row=1)

# Bind the Enter key to the search function
entry.bind("<Return>", search)

tk.mainloop()