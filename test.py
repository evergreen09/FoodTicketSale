from tkinter import *
from tkinter import StringVar
import tkinter.messagebox as msgbox
import tkinter.ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
from pandastable import Table, TableModel
import re
from datetime import datetime
import json

jsonName = 'sample.json'

# Define the font properties
font = Font(name='굴림체', size=9, bold=False)
nm_font = Font(name='맑은 고딕', size=11, bold=False)

class TicketNumber():
    def __init__(self, ticket_number):
        self._ticket_number = ticket_number

    @property
    def ticket_number(self):
        return self._ticket_number

    @ticket_number.setter
    def ticket_number(self, value):
        self._ticket_number = value
        ticket_var.set(self._ticket_number-1)  # Update the StringVar

class NonMember():
    def __init__(self, non_member_row):
        self.non_member_row = non_member_row

count = TicketNumber(1)
nm_row = NonMember(2)

def excel_loader(mont, day):
    # Retrieve month and date
    date = f'{day:02}'
    month = f'{mont:02}'

    # Set filename
    file_name = f'전체이용자(편집)_{month}_{date}'

    try:
        # Load the workbook and sheet
        workbook = load_workbook(f'{file_name}.xlsx')
        sheet = workbook['회원']
        print("Successfully Loaded Excel")
        return workbook, sheet, file_name
    except FileNotFoundError:
        msgbox.showerror("Error", "Excel file not found.")
        return None, None, None

def get_price(row):
    re_val = re.sub(r'\s+', '', sheet[f'C{row}'].value)
    # Assign appropriate price for users
    match re_val:
        case '기초생활수급권자' | '차상위(저소득)':
            sheet[f'D{row}'].value = 0
        case '기타' | '국가유공자':
            sheet[f'D{row}'].value = 1750
        case '일반':
            sheet[f'D{row}'].value = 3500
        case _:
            sheet[f'D{row}'].value = 9999
    # Apply font to the cell
    sheet[f'D{row}'].font = font

def altFood(user_name):
    names = []
    for name in names:
        if name == user_name:
            output("대체식")

def find_user(user_id, count):
    global df
    for cell in sheet['A']:
        if str(cell.value) == str(user_id): # Ensure both are compared as strings
            get_price(cell.row)
            sheet[f'E{cell.row}'].value = count
            sheet[f'E{cell.row}'].font = font
            return True
    return False

def warn():
    msgbox.showwarning("경고", "회원번호 불일치")

def output(text):
    msgbox.showinfo("Output", f'{text}')

# Load workbook and sheet
workbook, sheet, file_name = excel_loader(datetime.now().month, datetime.now().day)

tk = Tk()
tk.title("식권 판매 프로그램")
tk.geometry("800x800")

# Create the DataFrame viewer at the top
frame = Frame(tk)
frame.pack(fill=BOTH)

treeview = tkinter.ttk.Treeview(frame,
    columns=['UserID', 'Name', 'Status', 'price', 'ticketNumber'],
    displaycolumns=['UserID', 'Name', 'Status', 'price', 'ticketNumber'], show='headings')
treeview.pack()

treeview.column("UserID", width=100, anchor="center")
treeview.heading("UserID", text="회원번호", anchor="center")

treeview.column("Name", width=100, anchor="center")
treeview.heading("Name", text="이름", anchor="center")

treeview.column("Status", width=100, anchor="center")
treeview.heading("Status", text="생활구분", anchor="center")

treeview.column("price", width=100, anchor="center")
treeview.heading("price", text="가격", anchor="center")

treeview.column("ticketNumber", width=100, anchor="center")
treeview.heading("ticketNumber", text="식권", anchor="center")

# Create a frame to hold the entry and buttons horizontally
input_frame = Frame(tk)
input_frame.pack(pady=10)

entry = Entry(input_frame)
entry.grid(row=0, column=0, padx=5)

def search(event=None): # Add event parameter to handle binding
    id = entry.get() # Get the input as a string
    if find_user(id, count.ticket_number):
        count.ticket_number += 1
        workbook.save(f'{file_name}.xlsx')
        if count.ticket_number % 5 == 0:
            workbook.save(f'식권_백업\\{file_name}_백업{count.ticket_number}.xlsx')
            print("FILE BACKED UP")
    else:
        warn()
    entry.delete(0, 'end')

def delete():
    try:
        user_id = int(entry.get())
        for cell in sheet['A']:
            if cell.value == user_id:
                sheet[f'D{cell.row}'].value = None
                sheet[f'E{cell.row}'].value = None
                workbook.save(f'{file_name}.xlsx')
                output("삭제 완료")
                return
    except ValueError:
        warn()

def add_user(userID, name, status):
    try:
        print(userID, name, status)
        userID = int(userID)
        # Set Retrieved values to excel
        last_row = sheet.max_row + 1
        sheet[f'A{last_row}'].value = userID
        sheet[f'B{last_row}'].value = name
        sheet[f'C{last_row}'].value = status

        # Change font
        sheet[f'A{last_row}'].font = font
        sheet[f'B{last_row}'].font = font
        sheet[f'C{last_row}'].font = font

        # Align texts to center
        sheet[f'A{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'B{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'C{last_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        workbook.save(f'{file_name}_0.xlsx')
        output("신규 회원 추가 완료")
    except (IndexError, ValueError):
        warn()

def add_non_member(name, reason_for_visit):
    global df

    # Set Retrieved values to excel
    workbook['비회원'][f'A{nm_row.non_member_row}'].value = name
    workbook['비회원'][f'B{nm_row.non_member_row}'].value = reason_for_visit
    workbook['비회원'][f'C{nm_row.non_member_row}'].value = '3500'
    workbook['비회원'][f'D{nm_row.non_member_row}'].value = count.ticket_number

    # Change font
    workbook['비회원'][f'A{nm_row.non_member_row}'].font = nm_font
    workbook['비회원'][f'B{nm_row.non_member_row}'].font = nm_font
    workbook['비회원'][f'C{nm_row.non_member_row}'].font = nm_font
    workbook['비회원'][f'D{nm_row.non_member_row}'].font = nm_font

    workbook.save(f'{file_name}.xlsx')
    count.ticket_number += 1
    nm_row.non_member_row += 1

def jsonUpdate(userID, name, status, price):
    global jsonName
    new_data = {
            "user_id": userID,
            "user_name": name,
            "status": status,
            "price": price,
            "ticket_number": "1" 
        }
    print(userID)
    # treeview.insert('', 'end', text='', values=(f'{userID} {name} {status} {price} 1'))
    treeview.insert('', 'end', text='', values=(f'1 2 3 4'))
    print('inserted')
    # Open the JSON file and load the data
    with open(jsonName, 'r', encoding='utf-8') as file:
        data = json.load(file)

    # Append new data to the list
    data.append(new_data)

    # Write the updated data back to the file
    with open(jsonName, 'w', encoding='utf-8') as file:
        json.dump(data, file, indent=4, ensure_ascii=False)

def reset_ticket_number():
    reset_value = int(entry.get())
    count.ticket_number = reset_value

def save_file():
    dir = f'C:\\Users\\Administrator\\Desktop\\식권2024\\8월\\{file_name}.xlsx'
    workbook.save(dir)
    print("FILE SAVED SUCCESSFULLY!!!")

def addUserPopUp():
    addUserWindow = Toplevel(tk)
    addUserWindow.geometry('500x300')
    addUserWindow.title('신규회원추가')
    userID_text = Entry(addUserWindow)
    name_text = Entry(addUserWindow)
    status_text = Entry(addUserWindow)
    
    Label(addUserWindow, text='회원번호').grid(row=0, column=0, padx=5)
    userID_text.grid(row=0, column=1, padx=5)
    Label(addUserWindow, text='이름').grid(row=1, column=0, padx=5)
    name_text.grid(row=1, column=1, padx=5)
    Label(addUserWindow, text='생활구분').grid(row=2, column=0, padx=5)
    status_text.grid(row=2, column=1, padx=5)

    addUserButton = Button(addUserWindow, text='신규회원추가', command=lambda: add_user(userID_text.get(), name_text.get(), status_text.get()))
    addUserButton.grid(row=3, column=2, padx=5)

def nonUserPopUp():
    nonUserWindow = Toplevel(tk)
    nonUserWindow.geometry('500x300')
    nonUserWindow.title('비회원 판매')

    name_text = Entry(nonUserWindow)
    reason_for_visit_text = Entry(nonUserWindow)

    Label(nonUserWindow, text='이름').grid(row=0, column=0, padx=5)
    name_text.grid(row=0, column=1, padx=5)
    Label(nonUserWindow, text='방문용무').grid(row=1, column=0, padx=5)
    reason_for_visit_text.grid(row=1, column=1, padx=5)

    addNonUserButton = Button(nonUserWindow, text='비회원 판매', command=lambda: add_non_member(name_text.get(), reason_for_visit_text.get()))
    addNonUserButton.grid(row=2, column=2, padx=5)

def debug():
    print('Debug')
    jsonUpdate('15151', '홍길동', '일반', 3500)

# Create the StringVar after initializing the root window
ticket_var = StringVar()

search_button = Button(input_frame, text='검색', command=search)
search_button.grid(row=0, column=1, padx=5)

delete_button = Button(input_frame, text='환불', command=delete)
delete_button.grid(row=0, column=2, padx=5)

add_user_button = Button(input_frame, text='신규회원추가', command=addUserPopUp)
add_user_button.grid(row=0, column=3, padx=5)

reset_number = Button(input_frame, text='식권번호리셋', command=reset_ticket_number)
reset_number.grid(row=0, column=4, padx=5)

non_memeber_button = Button(input_frame, text='비회원', command=nonUserPopUp)
non_memeber_button.grid(row=0, column=5, padx=5)

save_button = Button(input_frame, text='저장', command=save_file)
save_button.grid(row=0, column=6, padx=5)

save_button = Button(input_frame, text='Debug', command=debug)
save_button.grid(row=0, column=7, padx=5)

total_tickets_sold = Label(input_frame, textvariable=ticket_var, font=('Helvetica', 25, 'bold'))
total_tickets_sold.grid(row=1)

# Bind the Enter key to the search function
entry.bind("<Return>", search)

tk.mainloop()


